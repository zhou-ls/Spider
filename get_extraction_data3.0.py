# -*- coding: utf-8 -*-
# @Time : 2020/4/12 22:39
# @Author : zls
# @File : get_extraction_data2.0.py
# @Software: PyCharm
import argparse
import re
from openpyxl import load_workbook


def find_sentence(en1, en2, drug_interaction):
    drug_interaction = drug_interaction.split('。')
    all_sentence = []  # 存一个药品说明书的所有句子 先分行后分句 以句号分句
    sentence = []  # 包含所有药物对句子的列表
    for line in drug_interaction:
        # if line:
        #     line = line.replace(' ', '').replace('\n', '').replace('\r', '')
        #     all_sentence.append(line)
        line = line.split('\n')
        for i in line:
            if i:
                i = i.replace(' ', '')  # 去除句子中的空格
                all_sentence.append(i)
    for sen in all_sentence:
        if en1 in sen and en2 in sen:
            sentence.append(sen)
    return sentence


def get_name(content):
    en1 = re.findall("【通用名称】 (.*)\n.*", content)
    return en1


def get_all_drug(disease_name):
    f = open(disease_name, 'r', encoding='utf-8')
    all_drug = []
    while True:
        drug_name = f.readline()
        if drug_name == '':
            break
        # print(type(drug_name[:-1]))
        all_drug.append(drug_name[:-1])
    f.close()
    return all_drug


def get_value_drug_name(src1, src2, ins, all_drug, all_disease, i):
    src_drug = []
    int_drug = []
    if ins and ("本品" or "本药品" or "本药") not in ins:
        if src1:
            for j in all_drug:
                drug1 = j.replace(' ', '')
                if drug1 in src1:
                    src_drug.append(drug1)  # 每种药品所含成分名称
            # print("i:", i)
            # print("src1:", src_drug)
            for k in all_disease:
                drug2 = k.replace(' ', '')
                if drug2 in ins:
                    int_drug.append(drug2)  # 每种药品相互作用的疾病名称
        elif src2:
            for j in all_drug:
                drug1 = j.replace(' ', '')
                if drug1 in src2:
                    src_drug.append(drug1)  # 每种药品所含成分名称
            # print("i:", i)
            # print("src2:", src_drug)
            for k in all_disease:
                drug2 = k.replace(' ', '')
                if drug2 in ins:
                    int_drug.append(drug2)  # 每种药品相互作用的疾病名称
    if ins and ("本品" or "本药品" or "本药") in ins:
        if src2:
            src_drug = get_name(src2)
            for j in all_disease:
                drug2 = j.replace(' ', '')
                # if drug2 in ins:
                #     int_drug.append(drug2)  # 每种药品相互作用的疾病名称
                if drug2 in ins and src1 and drug2 not in src1:
                    int_drug.append(drug2)  # 每种药品相互作用的疾病名称
                if drug2 in ins and src1 and drug2 in src1:
                    src_drug.append(drug2)
            # if "氨酚氯雷伪麻缓释片" in src_drug:
            #     print("i:", i)
            #     print("src2本品:", src_drug)
            #     print("int_drug:", int_drug)
    en1 = []
    en2 = []
    if int_drug and src_drug:
        src_drug = sorted(src_drug, key=lambda x: len(x), reverse=False)  # 按药品名称长度升序排列
        int_drug = sorted(int_drug, key=lambda y: len(y), reverse=False)  # 按药品名称长度升序排列
        # print("row:", i)
        # print("en1:", src_drug)
        # print("en2:", int_drug)
        for m in range(len(src_drug) - 1):
            flag = 0
            for n in range(m + 1, len(src_drug)):
                if src_drug[m] in src_drug[n]:
                    flag = 1
            if flag == 0:  # 如果src_drug[m]不是子字符串
                en1.append(src_drug[m])
        en1.append(src_drug[-1])
        for m in range(len(int_drug) - 1):
            flag = 0
            for n in range(m + 1, len(int_drug)):
                if int_drug[m] in int_drug[n]:
                    flag = 1
            if flag == 0:  # 如果int_drug[m]不是子字符串
                en2.append(int_drug[m])
        en2.append(int_drug[-1])  # 最后一个字符串长度最长，一定不是子字符串
        if "氨酚氯雷伪麻缓释片" in en1:
            print("row:", i)
            print("en1:", en1)
            print("en2:", en2)
    return en1, en2  # 去掉子字符串后的药品列表


def write_file(en1, en2, src2, ins, f, i):
    if en1 and en2:
        if ins and "本品" in ins:
            src_drug = get_name(src2)
            if src_drug[0] == "氨酚氯雷伪麻缓释片":
                print(src_drug)
            ins = ins.replace("本品", src_drug[0])
        if ins and "本药品" in ins:
            src_drug = get_name(src2)
            ins = ins.replace("本药品", src_drug[0])
        if ins and "本药" in ins:
            src_drug = get_name(src2)
            ins = ins.replace("本药", src_drug[0])
        for d1 in en1:
            for d2 in en2:
                if '无' not in d2 and '明确' not in d2 and d1 != d2 and d2 not in d1 and d1 not in d2:
                    sentences = find_sentence(d1, d2, ins)
                    for sentence in sentences:
                        sentence = sentence.replace(' ', '')
                        sentence = sentence.replace('　', '')
                        # f.write(d1 + '\t' + d2 + '\t' + sentence + '\n')
                        f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence + '\n')  # 带有药品说明书excel表的行号


                        # if len(sentence) > 2 and sentence[0].isdigit() and sentence[1].isdigit() == False:  # 如果是一个数字开头的，如1.
                        #     f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence[2:] + '\n')
                        # elif len(sentence) > 3 and sentence[0].isdigit() and sentence[1].isdigit():  # 如果是两个数字开头的 如12.
                        #     f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence[3:] + '\n')
                        # elif len(sentence) > 3 and (sentence[0] == '(' or sentence[0] == '（') and sentence[1].isdigit() and (sentence[2] == ')' or sentence[2] == '）' ):  # 如果是括号加一个数字开头如(1)
                        #     f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence[3:] + '\n')
                        # elif len(sentence) > 3 and (sentence[0] == '(' or sentence[0] == '（') and sentence[1].isdigit() and sentence[2].isdigit() and (sentence[3] == ')' or sentence[3] == '）'): # 如果是括号加两个个数字开头如(12)
                        #     f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence[4:] + '\n')
                        # elif sentence[0] == '•' or sentence[0] == '♦' or sentence[0] == ':' or sentence[0] == '：':
                        #     f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence[1:] + '\n')
                        # else:
                        #     f.write(str(i) + '\t' + d1 + '\t' + d2 + '\t' + sentence + '\n')


def get_exchange_data(f, filenames, all_drug, all_disease, columns):
    for filename in filenames:  # 说明书数量
        wb = load_workbook(filename)
        sheets = wb.sheetnames  # 获取所有工作表名
        for sheet in sheets:  # 每个工作表的名称
            ac_sheet = wb[sheet]
            for i in range(2, ac_sheet.max_row + 1):  # 每循环一次查找一个说明书的一张表
                if i % 2000 == 0:
                    print(f"已读取文件《{filename}》第{i}行数据")
                for column in columns:
                    ins = ac_sheet.cell(row=i, column=column).value
                    src1 = ac_sheet.cell(row=i, column=2).value  # 药品成分列
                    src2 = ac_sheet.cell(row=i, column=1).value  # 具体的药品名称列
                    if ins:
                        ins = ins.replace(' ', '')
                    en1, en2 = get_value_drug_name(src1, src2, ins, all_drug, all_disease, i)
                    write_file(en1, en2, src2, ins, f, i)  # i为行号


# 去重
def remove_same(result):
    f2 = open(result, 'r', encoding='utf-8')
    lines = []
    while True:
        line = f2.readline()
        if line == '':
            break
        lines.append(line[:-1])
    # lines = list(set(lines))
    print("提取出的药品实体+句子的数量：", len(lines))
    return lines


def sort_by_gbk(lines):
    """

    使用gbk编码对汉字进行排序
    :return gbk.    能gbk编码的
            no_gbk. 不能gbk编码

    """
    gbk = []
    no_gbk = []
    for i in lines:
        try:
            # 能gbk编码
            test = i.encode('gbk')
            gbk.append(i)
        except:
            # 不能gbk编码
            no_gbk.append(i)
    gbk = sorted(gbk, key=lambda x: x.encode('gbk'))
    return gbk, no_gbk


def get_train_data(gbk, no_gbk, result):
    f3 = open(result, 'w', encoding='utf-8')
    for i in gbk:
        f3.write(i + '\n')
    for i in no_gbk:
        f3.write(i + '\n')
    f3.close()


def main():
    parser = argparse.ArgumentParser(description="the path of drug and disease lists")
    parser.add_argument('-dis', '--disease', default='all_drug_name.txt')
    parser.add_argument('-drug', '--drug', default='all_drug_name.txt')
    parser.add_argument('-result', '--result', default='药品总的实体对.txt')
    # parser.add_argument('-col', '--columns', nargs="+", type=int, default=[7, 11, 12, 16])
    # parser.add_argument('-f', '--file', nargs="+", type=str, default=['药品说明书（爬取结果）.xlsx'])
    args = parser.parse_args()
    disease_name = args.disease
    drug_name = args.drug
    result = args.result
    # filenames = args.file  # 要提取的说明书名称列表
    filenames = ['原始药品说明书(去重).xlsx']  # 要提取的说明书名称列表
    f = open(result, 'w', encoding='utf-8')
    # columns = args.columns  # 提取说明书第几列的数据 7 适应症  11 禁忌  12  注意事项   16  药物相互作用
    columns = [15]  # 提取说明书第几列的数据 7 适应症  11 禁忌  12  注意事项   16  药物相互作用
    all_disease = get_all_drug(disease_name)
    all_drug = get_all_drug(drug_name)
    get_exchange_data(f, filenames, all_drug, all_disease, columns)
    f.close()
    lines = remove_same(result)
    gbk, no_gbk = sort_by_gbk(lines)
    get_train_data(gbk, no_gbk, result)


if __name__ == "__main__":
    main()

    sentences = []
    with open('药品总的实体对.txt', 'r', encoding='utf-8') as f:
        while True:
            con = f.readline()
            if con == '':
                break
            sentences.append(con[:-1])
    sens = list(set(sentences))
    print(len(sens))

