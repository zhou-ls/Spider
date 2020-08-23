from openpyxl import load_workbook

file = load_workbook(r'test.xlsx')
active_sheet = file["Sheet1"]
for i in range(1, active_sheet.max_row + 1):  # 读行
    for j in range(1, active_sheet.max_column + 1):  # 读列
        value = active_sheet.cell(row=i, column=j).value
        if value is not None:
            print(value)
