# -*- coding: utf-8 -*-
# @Time : 2020/10/28 9:22
# @Author : zls
# @File : 120ask_spider.py
from pyquery import PyQuery as pq
from openpyxl import Workbook
import requests


headers = {'Content-Encoding': 'gzip',
           'Content-Type': 'text/html; charset=UTF-8',
           'Date': 'Wed, 28 Oct 2020 01:22:31 GMT',
           'Vary': 'Accept-Encoding',
           'X-Cache': 'bypass',
           'X-Via-JSL': '16d00f5,-',
           'Yii-server': '181',
           'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.80 Safari/537.36'}


def get_html(url):
    n = 0
    try:
        response = requests.get(url, headers=headers, timeout=3)
        response.encoding = response.apparent_encoding
        html = response.text
        return html
    except ConnectionError:
        n += 1
        if n < 4:
            print('连接超时，正在重新爬取数据......')
            return get_html(url)
        else:
            return None


def get_url():
    # 中西药品：https://yp.120ask.com/search/0-0-1--0-0-0-0.html
    zhongxi = []
    for i in range(1, 101):
        url = f'https://yp.120ask.com/search/0-0-{i}--0-0-0-0.html'
        zhongxi.append(url)

    # 家庭常备：https://yp.120ask.com/search/36-0-99--0-0-0-0.html
    jiating = []
    for i in range(1, 101):
        url = f'https://yp.120ask.com/search/36-0-{i}--0-0-0-0.html'
        jiating.append(url)

    # 男科用药：https://yp.120ask.com/search/9-0-1--0-0-0-0.html
    man = []
    for i in range(1, 101):
        url = f'https://yp.120ask.com/search/9-0-{i}--0-0-0-0.html'
        man.append(url)

    # 母婴用药：https://yp.120ask.com/search/11-0-1--0-0-0-0.html
    mother = []
    for i in range(1, 101):
        url = f'https://yp.120ask.com/search/11-0-{i}--0-0-0-0.html'
        mother.append(url)

    # 中老年用药：https://yp.120ask.com/search/12-0-1--0-0-0-0.html
    old = []
    for i in range(1, 101):
        url = f'https://yp.120ask.com/search/12-0-{i}--0-0-0-0.html'
        old.append(url)
    return zhongxi, jiating, man, mother, old


# 查找每个具体药品的url
def search(row, url, titleName, currentSheet):
    t = []  # title
    c = []  # the content of the title
    print(url)
    html = get_html(url)
    if html:
        doc = pq(html)
        urls1 = doc('.Sort-list.Drug-store ul li div i a').items()
        for detail in urls1:
            manual = detail.attr('href').replace('detail', 'manual')
            detail_url = 'https://yp.120ask.com' + manual
            # print(detail_url)
            html = get_html(detail_url)
            if html:
                doc = pq(html)
                title = doc('.cont-Drug-details .cont-2.tab-dm-2 .table .tabrow .td').items()
                content = doc('.cont-Drug-details .cont-2.tab-dm-2 .table .tabrow .td-details').items()
                for tit in title:
                    # print(tit.text())
                    t.append(tit.text())
                    if tit.text() not in titleName:
                        titleName.append(tit.text())
                for con in content:
                    # print(con.text())
                    c.append(con.text())
                dic = dict(zip(t, c)).items()
                print(dic)
                # print(titleName)
                for k, v in dic:
                    currentSheet.cell(row=row, column=titleName.index(k) + 1, value=v)
                print(f'已爬取{row - 1}条')
                row = row + 1
        return row


def save_file(wb, currentSheet, file_name, url_list):
    titleName = []
    row = 2
    for page_url in url_list:  # page_url为每一页的url
        row = search(row, page_url, titleName, currentSheet)
    for i in range(1, len(titleName) + 1):
        currentSheet.cell(row=1, column=i, value=titleName[i - 1])
    print(titleName)
    wb.save(file_name)
    print('爬取完毕！！！')


def creat_excel():
    wb = Workbook()
    wb.create_sheet(index=0, title="Sheet1")
    currentSheet = wb.active
    return wb, currentSheet


def run_spider():
    """
    分步爬取数据，生成五个文件：中西药品；家庭常备；男科用药；母婴用药；中老年用药
    :return:
    """
    zhongxi, jiating, man, mother, old = get_url()
    # 创建excel文档,中西药品
    wb, currentSheet = creat_excel()
    file_name1 = '120ask中西药品说明书.xlsx'
    save_file(wb, currentSheet, file_name1, zhongxi)

    # 创建excel文档, 家庭常备
    wb, currentSheet = creat_excel()
    file_name2 = '120ask家庭常备说明书.xlsx'
    save_file(wb, currentSheet, file_name2, jiating)

    # 创建excel文档男科用药, 男科用药
    wb, currentSheet = creat_excel()
    file_name3 = '120ask男科用药说明书.xlsx'
    save_file(wb, currentSheet, file_name3, man)

    # 创建excel文档,母婴用药
    wb, currentSheet = creat_excel()
    file_name4 = '120ask母婴用药说明书.xlsx'
    save_file(wb, currentSheet, file_name4, mother)

    # 创建excel文档,中老年用药
    wb, currentSheet = creat_excel()
    file_name5 = '120ask中老年用药说明书.xlsx'
    save_file(wb, currentSheet, file_name5, old)


if __name__ == '__main__':
    run_spider()
