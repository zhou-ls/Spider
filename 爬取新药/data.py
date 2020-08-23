from pyquery import PyQuery as pq
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
import requests
from time import sleep

# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active

# # 创建chrome浏览器驱动，无头模式
# chrome_options = Options()
# # chrome_options.add_argument('--headless')
# chrome_options.add_argument("--start-maximized")
# browser = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)

# 维持会话
session = requests.Session()
session.keep_alive = False


def get_html(url):
    try:
        response = session.get(url)
        print('aaaaaaaaaaaaaaaaaaaaaaa')
        # 解决中文乱码
        response.encoding = response.apparent_encoding
        print('hhhhhhhhhhhhhhhhhhhhhh')
        html = response.text
        return html
    except ConnectionError:
        # print('连接超时，正在重新爬取数据......')
        # get_html(url)
        return None


# # 如果网络超时，则重新爬取数据
# def get_html(url):
#     try:
#         browser.get(url)
#         html = browser.page_source
#         return html
#     except TimeoutException:
#         print('网络超时，正在尝试重新爬取数据......')
#         get_html(url)


# 查找每个具体药品的url
def search(row, url, titleName):
    row = row + 1
    t = []
    c = []
    print(url)
    html = get_html(url)
    print(row)
    if html:
        doc = pq(html)
        ins = doc('.pro_tab.mart25.white .p-detailcon.changeModel .p-con.Swtab .margin15.prodetab_p')
        title = doc('.pro_tab.mart25.white .p-detailcon.changeModel .p-con.Swtab .margin15.prodetab_p .ifolable').items()
        ins.find('p.ifolable').remove()
        content = ins.find('p').items()
        for tit in title:
            # print(tit.text())
            t.append(tit.text())
            if tit.text() not in titleName:
                titleName.append(tit.text())
            # print('****************************************************************************')
        for con in content:
            # print(con.text())
            c.append(con.text())
        dic = dict(zip(t, c)).items()
        # print(titleName)
        for k, v in dic:
            for j in range(1, len(list(zip(t, c))) + 1):
                currentSheet.cell(row=row, column=titleName.index(k) + 1, value=v)
        # print(f'已爬取{row - 1}条')
    return row


def main():
    titleName = []
    row = 1
    f = open('all_url.txt', 'r', encoding='utf-8')
    while True:
        url = f.readline()
        if url == '':
            break
        url1 = url[:-1]
        row = search(row, url1, titleName)
    wb.save('新特网药品说明书.xlsx')
    for i in range(1, len(titleName) + 1):
        currentSheet.cell(row=1, column=i, value=titleName[i - 1])
    print(titleName)
    wb.save('新特网药品说明书.xlsx')
    print('爬取完毕！！！')


if __name__ == '__main__':
    main()
