# -*- coding: utf-8 -*-
# @Time : 2020/8/16 0:20
# @Author : zls
# @File : all_url.py
# @Software: PyCharm
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from pyquery import PyQuery as pq
from openpyxl import Workbook
import requests
import time


# # 创建excel文档
# wb = Workbook()
# wb.create_sheet(index=0, title="Sheet1")
# currentSheet = wb.active


# 创建chrome浏览器驱动，无头模式
chrome_options = Options()
# chrome_options.add_argument('--headless')
chrome_options.add_argument("--start-maximized")
driver = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)

# 维持会话
session = requests.Session()
session.keep_alive = False

f = open('all_url.txt', 'w', encoding='utf-8')


def driver_get(url):
    try:
        driver.get(url)
    except TimeoutException:
        driver_get(url)


def get_html(url):
    try:
        response = session.get(url)
        # 解决中文乱码
        response.encoding = response.apparent_encoding
        html = response.text
        return html
    except ConnectionError:
        print('连接超时，正在重新爬取数据......')
        get_html(url)


def parent_url(url):
    html = get_html(url)
    doc = pq(html)
    urls = []
    allclass = doc('.container.gray01.topfixed .menu .nav dd a').items()
    for url in allclass:
        urls.append(url.attr('href'))
    return urls


#  需要通过滑动来动态渲染的url
def slide(url):
    # url = "http://m.xinyao.com.cn/druglist/366"
    # 加载界面
    driver_get(url)
    time.sleep(3)

    # 获取页面初始高度
    js = "return action=document.body.scrollHeight"
    height = driver.execute_script(js)

    # 将滚动条调整至页面底部
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    time.sleep(10)

    # 定义初始时间戳（秒）
    t1 = int(time.time())

    # 定义循环标识，用于终止while循环
    status = True

    # 重试次数
    num = 0

    while status:
        # 获取当前时间戳（秒）
        t2 = int(time.time())
        # 判断时间初始时间戳和当前时间戳相差是否大于30秒，小于30秒则下拉滚动条
        if t2 - t1 < 30:
            new_height = driver.execute_script(js)
            if new_height > height:
                time.sleep(1)
                driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
                # 重置初始页面高度
                height = new_height
                # 重置初始时间戳，重新计时
                t1 = int(time.time())
        elif num < 4:  # 当超过30秒页面高度仍然没有更新时，进入重试逻辑，重试3次，每次等待30秒
            time.sleep(5)
            num = num + 1
        else:  # 超时并超过重试次数，程序结束跳出循环，并认为页面已经加载完毕！
            print("滚动条已经处于当前页面最下方！")
            status = False
            # # 滚动条调整至页面顶部
            # driver.execute_script('window.scrollTo(0, 0)')
            break

    time.sleep(10)
    # 打印页面源码
    html = driver.page_source
    doc = pq(html)
    # 每个大类所有具体药品的url,为一个class
    drug_urls = doc('.container.gray01.topfixed .wr .pro_item a').items()
    return drug_urls


# 查找每个具体药品的url
def search(row, drug_urls, titleName):
    if drug_urls:
        for drug_url in drug_urls:
            t = []
            c = []
            row = row + 1
            url = drug_url.attr('href')
            print(url)
            f.write(url + '\n')
    #         html = get_html(url)
    #         doc = pq(html)
    #         ins = doc('.pro_tab.mart25.white .p-detailcon.changeModel .p-con.Swtab .margin15.prodetab_p')
    #         title = doc('.pro_tab.mart25.white .p-detailcon.changeModel .p-con.Swtab .margin15.prodetab_p .ifolable').items()
    #         ins.find('p.ifolable').remove()
    #         content = ins.find('p').items()
    #         for tit in title:
    #             # print(tit.text())
    #             t.append(tit.text())
    #             if tit.text() not in titleName:
    #                 titleName.append(tit.text())
    #             # print('****************************************************************************')
    #         for con in content:
    #             # print(con.text())
    #             c.append(con.text())
    #         dic = dict(zip(t, c)).items()
    #         # print(titleName)
    #         for k, v in dic:
    #             for j in range(1, len(list(zip(t, c))) + 1):
    #                 currentSheet.cell(row=row, column=titleName.index(k) + 1, value=v)
    #         # print(f'已爬取{row - 1}条')
    return row


def main():
    titleName = []
    base_url ='http://m.xinyao.com.cn/allclass.html'
    urls = parent_url(base_url)
    row = 1
    for url in urls:
        print(url)
        # if url == 'http://m.xinyao.com.cn/druglist/1053':
        drug_urls = slide(url)
        row = search(row, drug_urls, titleName)
        print(f'-------------------已爬取{row - 1}条数据-----------------------')
    #     wb.save('新特网药品说明书.xlsx')
    # for i in range(1, len(titleName) + 1):
    #     currentSheet.cell(row=1, column=i, value=titleName[i - 1])
    # print(titleName)
    # wb.save('新特网药品说明书.xlsx')
    # print('爬取完毕！！！')


if __name__ == '__main__':
    main()

