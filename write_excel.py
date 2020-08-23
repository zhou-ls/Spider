import openpyxl

name = ["test1", "test2", "test3", "test4"]


# 创建excel并写入数据
def create_sheet_cell():
    wb = openpyxl.Workbook()    # for n in range(nSheet):
    wb.create_sheet(index=0, title="Sheet1")
    currentSheet = wb.active
    for i in range(1, 6):
        for j in range(1, len(name) + 1):
            currentSheet.cell(row=i, column=j, value=name[j - 1])
    wb.save("test.xlsx")


if __name__ == "__main__":
    # nSheet = int(input("输入电子表格的个数（整数）："))
    # nCell = int(input("输入乘法表的最大值（整数）："))
    create_sheet_cell()
