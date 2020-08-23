# -*- coding: utf-8 -*-
# @Time : 2020/8/11 8:25
# @Author : zls
# @File : jianke.py
# @Software: PyCharm
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook
import requests
import re

titltname = ["药品名称", "主要成份", "性 状", "适应症/功能主治", "规格型号", "用法用量", "不良反应", "禁 忌", "注意事项", "药物相互作用", "贮 藏", "包 装", "有 效 期",
             "执行标准", "批准文号", "生产企业"]

# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active
for i in range(1, len(titltname) + 1):
    currentSheet.cell(row=1, column=i, value=titltname[i - 1])

# excel文档追加
# file = load_workbook(r'健客网药品说明书.xlsx')
# currentSheet = file['Sheet1']

# 维持会话
session = requests.Session()
session.keep_alive = False


def get_html(url):
    response = session.get(url)
    # 解决中文乱码
    response.encoding = response.apparent_encoding
    html = response.text
    return html


def get_instruction(html, drug_url, row):
    doc = pq(html)
    product_urls = doc('.main .list-right .pro-list .pro-area .pro-con li .lihover .pro-botxt p a').items()
    next_pages = doc('.main .list-right .pages a').items()
    for product_url in product_urls:
        try:
            row = row + 1
            # 得到每个大类下每个具体药品的url
            if drug_url == 'https://www.jianke.com/Category/1315.html':  # 第六个大类的URL与众不同
                product_url = 'https://www.jianke.com' + product_url.attr('href')
            else:
                product_url = 'https:' + product_url.attr('href')
            # 药品详情页
            html = get_html(product_url)
            doc = pq(html)
            # ins_title = doc('.main .prod_main #pr_main_right .decora.tabs_mod.qiehuan .decora_cons .contmdiv .bigfont p em').items()
            ins_content = doc(
                '.main .prod_main #pr_main_right .decora.tabs_mod.qiehuan .decora_cons .contmdiv .bigfont p').text()
            # 使最后一个通过正则也匹配到
            ins_content = ins_content + ' 【'
            # 内容无法用html标签解析，用正则匹配出来
            nc = re.findall('【药品名称】 (.*?) 【.*', ins_content)
            # print(nc)
            cf = re.findall('【主要成份】 (.*?) 【.*', ins_content)
            xz = re.findall('【性 状】 (.*?) 【.*', ins_content)
            sg = re.findall('【适应症/功能主治】 (.*?) 【.*', ins_content)
            gg = re.findall('【规格型号】 (.*?) 【.*', ins_content)
            yy = re.findall('【用法用量】 (.*?) 【.*', ins_content)
            bl = re.findall('【不良反应】 (.*?) 【.*', ins_content)
            jj = re.findall('【禁 忌】 (.*?) 【.*', ins_content)
            zy = re.findall('【注意事项】 (.*?) 【.*', ins_content)
            xh = re.findall('【药物相互作用】 (.*?) 【.*', ins_content)
            cc = re.findall('【贮 藏】 (.*?) 【.*', ins_content)
            bz = re.findall('【包 装】 (.*?) 【.*', ins_content)
            yx = re.findall('【有 效 期】 (.*?) 【.*', ins_content)
            zx = re.findall('【执行标准】 (.*?) 【.*', ins_content)
            wh = re.findall('【批准文号】 (.*?) 【.*', ins_content)
            qy = re.findall('【生产企业】 (.*?) 【.*', ins_content)
            currentSheet.cell(row=row, column=1, value=''.join(nc))
            currentSheet.cell(row=row, column=2, value=''.join(cf))
            currentSheet.cell(row=row, column=3, value=''.join(xz))
            currentSheet.cell(row=row, column=4, value=''.join(sg))
            currentSheet.cell(row=row, column=5, value=''.join(gg))
            currentSheet.cell(row=row, column=6, value=''.join(yy))
            currentSheet.cell(row=row, column=7, value=''.join(bl))
            currentSheet.cell(row=row, column=8, value=''.join(jj))
            currentSheet.cell(row=row, column=9, value=''.join(zy))
            currentSheet.cell(row=row, column=10, value=''.join(xh))
            currentSheet.cell(row=row, column=11, value=''.join(cc))
            currentSheet.cell(row=row, column=12, value=''.join(bz))
            currentSheet.cell(row=row, column=13, value=''.join(yx))
            currentSheet.cell(row=row, column=14, value=''.join(zx))
            currentSheet.cell(row=row, column=15, value=''.join(wh))
            currentSheet.cell(row=row, column=16, value=''.join(qy))
            # print(f'-------------------已爬取{row - 1}条数据-----------------------')
        except:
            # print(product_url.attr('href'))
            # 有个别药品的网址与众不同，可能那边服务器有问题，无响应
            print("一个网址出现错误.....")
            continue
    #  翻页查找
    if next_pages:
        for page in next_pages:
            if page.find('i').text() == '>' and page.attr('href') != "javascript:void(0);":
                html = get_html('https:' + page.attr('href'))
                row = get_instruction(html, drug_url, row)
    return row


def main():
    # 网页首页地址
    html = get_html('https://www.jianke.com/')
    doc = pq(html)
    # 首页所有大类（男科用药、心血管科、糖尿病科、风湿骨科......）的地址的集合
    classifies = doc(
        '.jkn_nav .jkn_navigation .jkn_nav_l .jnk_allsort.jnk_allsorthover .mc .item .jnk_a_dl dt a').items()
    # excel表的行数
    row = 1
    for url in classifies:
        # 得到每个大类的url
        drug_url = 'https:' + url.attr('href')
        print('正在爬取的药品大类地址：', drug_url)
        html = get_html(drug_url)
        row = get_instruction(html, drug_url, row)
        print(f'-------------------已爬取{row - 1}条数据-----------------------')
        wb.save('健客网药品说明书.xlsx')


if __name__ == '__main__':
    main()
