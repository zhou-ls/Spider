# -*- coding: utf-8 -*-
# @Time : 2020/8/10 16:54
# @Author : zls
# @File : get_ins_data.py
# @Software: PyCharm
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
from random import randint
from time import sleep
from openpyxl import Workbook
import re


titltname = ["药品名称", "主要成份", "性 状", "适应症/功能主治", "规格型号", "用法用量", "不良反应", "禁 忌", "注意事项", "药物相互作用", "贮 藏",
             "包 装", "有 效 期", "执行标准", "批准文号", "生产企业"]

# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active
for i in range(1, len(titltname) + 1):
    currentSheet.cell(row=1, column=i, value=titltname[i - 1])


def browser_get(browser, url):
    try:
        browser.get(url)
    except TimeoutException:
        browser_get(browser, url)


def get_instruction(browser, html, drug_url,row):
    doc = pq(html)
    product_urls = doc('.main .list-right .pro-list .pro-area .pro-con li .lihover .pro-botxt p a').items()
    next_pages = doc('.main .list-right .pages a').items()
    for product_url in product_urls:
        row = row + 1
        # 得到每个大类下每个具体药品的url
        if drug_url == 'https://www.jianke.com/Category/1315.html':
            # print('https://www.jianke.com' + product_url.attr('href'))
            product_url = 'https://www.jianke.com' + product_url.attr('href')
        else:
            # print('https:' + product_url.attr('href'))
            product_url = 'https:' + product_url.attr('href')
        browser_get(browser, product_url)
        # 药品详情页
        html = browser.page_source
        doc = pq(html)
        # ins_title = doc('.main .prod_main #pr_main_right .decora.tabs_mod.qiehuan .decora_cons .contmdiv .bigfont p em').items()
        ins_content = doc('.main .prod_main #pr_main_right .decora.tabs_mod.qiehuan .decora_cons .contmdiv .bigfont p').text()
        # 使最后一个也匹配到
        ins_content = ins_content + ' 【'
        # print(ins_content)
        nc = re.findall('【药品名称】 (.*?) 【.*', ins_content)
        print(nc)
        cf = re.findall('【主要成份】 (.*?) 【.*', ins_content)
        xz = re.findall('【性 状】 (.*?) 【.*', ins_content)
        sg = re.findall('【适应症/功能主治】 (.*?) 【.*', ins_content)
        gg = re.findall('【规格型号】 (.*?) 【.*', ins_content)
        yy = re.findall('【用法用量】 (.*?) 【.*', ins_content)
        bl = re.findall('【不良反应】 (.*?) 【.*', ins_content)
        jj = re.findall('【禁 忌】 (.*?) 【.*', ins_content)
        zy = re.findall('【注意事项】 (.*?) 【.*', ins_content)
        xh = re.findall('【药物相互作用】 (.*?) 【.*', ins_content)
        cc = re.findall('【贮 藏】 (.*?) 【.*', ins_content)
        bz = re.findall('【包 装】 (.*?) 【.*', ins_content)
        yx = re.findall('【有 效 期】 (.*?) 【.*', ins_content)
        zx = re.findall('【执行标准】 (.*?) 【.*', ins_content)
        wh = re.findall('【批准文号】 (.*?) 【.*', ins_content)
        qy = re.findall('【生产企业】 (.*?) 【.*', ins_content)
        currentSheet.cell(row=row, column=1, value=''.join(nc))
        currentSheet.cell(row=row, column=2, value=''.join(cf))
        currentSheet.cell(row=row, column=3, value=''.join(xz))
        currentSheet.cell(row=row, column=4, value=''.join(sg))
        currentSheet.cell(row=row, column=5, value=''.join(gg))
        currentSheet.cell(row=row, column=6, value=''.join(yy))
        currentSheet.cell(row=row, column=7, value=''.join(bl))
        currentSheet.cell(row=row, column=8, value=''.join(jj))
        currentSheet.cell(row=row, column=9, value=''.join(zy))
        currentSheet.cell(row=row, column=10, value=''.join(xh))
        currentSheet.cell(row=row, column=11, value=''.join(cc))
        currentSheet.cell(row=row, column=12, value=''.join(bz))
        currentSheet.cell(row=row, column=13, value=''.join(yx))
        currentSheet.cell(row=row, column=14, value=''.join(zx))
        currentSheet.cell(row=row, column=15, value=''.join(wh))
        currentSheet.cell(row=row, column=16, value=''.join(qy))
        print(f'-------------------已爬取{row - 1}条数据-----------------------')
        wb.save('药品说明书.xlsx')
    if next_pages:
        for page in next_pages:
            if page.find('i').text() == '>' and page.attr('href') != "javascript:void(0);":
                # print('https:' + page.attr('href'))
                browser_get(browser, 'https:' + page.attr('href'))
                # sleep(60)
                html = browser.page_source
                row = get_instruction(browser, html, drug_url, row)
    return row


def main():
    browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe")
    wait = WebDriverWait(browser, 10)
    browser.get('https://www.jianke.com/')
    html = browser.page_source
    doc = pq(html)
    classifies = doc(
        '.jkn_nav .jkn_navigation .jkn_nav_l .jnk_allsort.jnk_allsorthover .mc .item .jnk_a_dl dt a').items()
    row = 1
    for url in classifies:
        # 得到每个大类的url
        print('======================================================================')
        # print('https:' + url.attr('href'))
        drug_url = 'https:' + url.attr('href')
        browser_get(browser, drug_url)
        html = browser.page_source
        row = get_instruction(browser, html, drug_url, row)


if __name__ == '__main__':
    main()
