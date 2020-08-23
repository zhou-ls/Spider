import re
from time import sleep
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook
from random import randint

# chrome_options = webdriver.ChromeOptions()
# chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])  # 以键值对的形式加入参数,隐藏selenium防一定程度的反爬
# chrome_options.add_argument('headless')
# browser = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)
browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe")
wait = WebDriverWait(browser, 10)

titltname = ["药品通用名(成分)", "药品名称", "成份", "所属类别", "性状", "作用类别", "适应症", "规格", "用法用量", "不良反应", "禁忌", "注意事项",
             "孕妇及哺乳期妇女用药", "儿童用药", "老年用药", "药物相互作用", "药物过量", "临床试验", "药理毒理", "药代动力学", "贮藏", "包装",
             "有效期", "执行标准", "批准文号", "进口药品注册证号", "委托方企业", "生产企业", "包装企业", "妊娠分级", "哺乳期分级", "功能主治", "资料来源",
             "警示语", "药理作用"
             ]

# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active
for i in range(1, len(titltname) + 1):
    currentSheet.cell(row=1, column=i, value=titltname[i - 1])

# 打开excel文档追加爬取
# wb = load_workbook("药品说明书(全).xlsx")
# currentSheet = wb["Sheet1"]


# 如果要读取多个药品，则反复调用search()
def search(drugName, row, space):
    try:
        browser.get('http://drugs.medlive.cn/drugref/index.do')
        input1 = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#q')))
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#drugSearchForm > input')))
        # drugName = input("输入需要查找的药品名称:  ")
        input1.send_keys(drugName)
        # input1.send_keys('阿瑞匹坦')
        submit.click()
        html = browser.page_source
        doc = pq(html)
        num = doc('.search-content .box-list .fenye').find('span').text()
        i = 0
        if num:
            drug_number = re.findall(".*页(.*)条.*", num)
            print("搜索出与输入的药品名相关的药物有{n}种：".format(n=drug_number[0]))
        row1 = get_instruction(drugName, row, i, space)
        print('当前爬取到的说明书数量为row1  ===', row1 - 1)
        print('============================================================================================')
        return row1
    except TimeoutException:
        print("当前网络状况较差，爬取时间超时，正在尝试重新爬取该药品...")
        return search(drugName, row, space)


def get_instruction(drugName, row, i, space):
    print("正在获取网页数据......")
    # wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#info-content .info-left')))
    html = browser.page_source
    doc = pq(html)
    medince_name = doc('.search-content .box-list .box1 .medince-name').text()
    medince_urls = doc('.search-content .box-list .box1').items()
    page = []
    page_url = []
    pages = doc('.search-content .box-list .fenye .grey').items()
    page_urls = doc('.search-content .box-list .fenye .grey').items()
    if len(medince_name.strip().split()) >= 1 and "适应症" not in medince_name:
        print(f"搜索出当前页面的药物有{str(len(medince_name.strip().split()))}种：{medince_name}")
        for medince_url in medince_urls:
            print("第{n}种说明书如下：".format(n=i + 1))
            i = i + 1
            sleep(randint(5, 10))  # 防止访问过于频繁，网站出现验证码
            url = medince_url.find('.medince-name a').attr('href')
            # print(url)
            browser.get('http://drugs.medlive.cn/' + url)
            html = browser.page_source
            doc = pq(html)
            row = row + 1
            print('row  ===', row)
            row = get_ins(row, doc)
            currentSheet.cell(row=row, column=1, value=drugName)
            print('============================================================================================')
        if pages and page_urls:
            for p in pages:
                page.append(p.text())
            for url in page_urls:
                page_url.append(url.find('a').attr('href'))
            dic = dict(zip(page, page_url))
            # print(dic)
            if '>' in dic:
                browser.get('http://drugs.medlive.cn/' + dic['>'])
                row = get_instruction(drugName, row, i, space)
        return row
    elif "适应症" in medince_name:
        space.append(drugName)
        print(f"当前有{len(list(set(space)))}种空白的药品：{list(set(space))}")
        print('============================================================================================')
        return row
    else:
        print(f"搜索出与输入的药品名相关的药物有1种：{medince_name}")
        row = row + 1
        print('row  ===', row)
        row = get_ins(row, doc)
        print('============================================================================================')
        currentSheet.cell(row=row, column=1, value=drugName)
        return row


def get_ins(row, doc):
    titles = doc('.info-content .info-left .title').items()
    informations = doc('.info-content .info-left .more-infomation').items()
    tit = []
    inf = []
    for title in titles:
        # print("名称-------", pq(title)('.title').text())
        tit.append(pq(title)('.title').text())
    for information in informations:
        # print("内容-------", pq(information)('.more-infomation').text())
        inf.append(pq(information)('.more-infomation').text())
    # print(dict(zip(tit, inf)))
    for k, v in dict(zip(tit, inf)).items():
        if k[:-1] in titltname and k[:-1] == titltname[titltname.index(k[:-1])]:
            for j in range(1, len(list(zip(tit, inf))) + 1):
                currentSheet.cell(row=row, column=titltname.index(k[:-1]) + 1, value=v)
        if k == "药物相互作用：":
            print(v)
    return row


def main():
    f = open(r'../origin_data/all_drug_name.txt', 'r', encoding='utf-8')
    f1 = open(r'../result_data/没有相关说明书的药品名称.txt', 'w', encoding='utf-8')
    space = []
    n = 0
    row = 1
    while True:
        drugName = f.readline()
        n = n + 1
        if drugName == '':
            break
        print(f"正在搜索第{n}个药品:   {drugName[:-1]}")
        row = search(drugName[:-1], row, space)
        wb.save("药品说明书.xlsx")
        sleep(randint(5, 10))  # 防止访问过于频繁，网站出现验证码
    for i in space:
        f1.write(i + '\n')


if __name__ == '__main__':
    main()
