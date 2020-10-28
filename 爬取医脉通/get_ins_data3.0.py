import requests
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
from random import randint, randrange
from time import sleep
from openpyxl import Workbook


"""
爬取相关的药品打开相关的注释，相关药品需要爬取的字段根据药师实际需要添加需要爬取的字段

需要爬取的字段要与网站上面的字段保持一致（除去网站上面的字段后面的冒号）

"""
PROXY_POOL_URL = 'http://localhost:5555/random'
options = webdriver.ChromeOptions()
# 西药和中成药爬取的条目字段
titltname = ["药品名称", "成份", "所属类别", "性状", "作用类别", "适应症", "规格", "用法用量", "不良反应", "禁忌", "注意事项",
             "孕妇及哺乳期妇女用药", "儿童用药", "老年用药", "药物相互作用", "药物过量", "临床试验", "药理毒理", "药代动力学", "贮藏", "包装",
             "有效期", "执行标准", "批准文号", "进口药品注册证号", "委托方企业", "生产企业", "包装企业", "妊娠分级", "哺乳期分级", "功能主治", "资料来源",
             "警示语", "药理作用"
             ]

# # 中药方剂爬取的条目字段
# titltname = ["药品名称", "主要成份", "所属类别", "组成", "用法", "功用", "主治", "配伍特点", "鉴别", "辩证要点", "加减变化", "现代运用", "使用注意"
#              ]

# # 中药中药材爬取的条目字段
# titltname = ["药品名称", "所属类别", "古籍出处", "基原", "主要产地", "采集", "炮制", "性状", "贮藏", "药性", "归经", "应用", "功效", "用法用量", "古籍摘要",
#              "化学成分", "药理作用"
#              ]


# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active
for i in range(1, len(titltname) + 1):
    currentSheet.cell(row=1, column=i, value=titltname[i - 1])


# 设置等待时间，防止访问速度过快，网站出现验证码，试过相关的验证码的识别方法，识别效果很差，
# 没有好的解决方法，如果出现验证码，目前只有人工输入验证码
def wait_yzm():
    # time = randrange(5, 25, 7)
    # time = randint(50, 60)
    # print('等待时间:', time, 's')
    sleep(0)


def get_proxy():
    """
    从代理池获取代理
    :return:
    """
    try:
        response = requests.get(PROXY_POOL_URL)
        if response.status_code == 200:
            print('Get Proxy', response.text)
            return response.text
        return None
    except requests.ConnectionError:
        return None


# 如果网络超时，则重新爬取数据
def browser_get(browser, url):
    try:
        browser.get(url)
    except TimeoutException:
        print('网络超时，正在尝试重新爬取数据......')
        browser_get(browser, url)


def login(browser):
    # chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])  # 以键值对的形式加入参数,隐藏selenium防一定程度的反爬
    # chrome_options.add_argument('headless')
    # browser = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)
    wait = WebDriverWait(browser, 10)
    browser.get('http://www.medlive.cn/auth/login')
    submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                    'body > div.large-box.login-box > div.page-main > div > div.main-right > div.login-rightTab.clearfix.qr-goTab > div.rightTab-L')))
    submit.click()
    input1 = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#username')))
    # 医脉通的用户名
    input1.send_keys('18707117829')
    input2 = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#password')))
    # 当前用户名的密码
    input2.send_keys('u8gbv6')
    submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#loginsubmit')))
    submit.click()


def get_instruction(browser, html, row, i):
    doc = pq(html)
    medince_name = doc('.search-content .box-list .box1 .medince-name').text()
    medince_urls = doc('.search-content .box-list .box1').items()
    page = []
    page_url = []
    pages = doc('.search-content .fenye .grey').items()
    page_urls = doc('.search-content .fenye .grey').items()
    if len(medince_name.strip().split()) >= 1 and "适应症" not in medince_name:
        print(f"搜索出当前页面的药物为：  {medince_name}")
        # print(f"搜索出当前页面的药物有{str(len(medince_name.strip().split()) // 3)}种：{medince_name}")
        for medince_url in medince_urls:
            print("第{n}种说明书如下：".format(n=i + 1))
            i = i + 1
            wait_yzm()
            url2 = 'http://drugs.medlive.cn/' + medince_url.find('.medince-name a').attr('href')
            # browser.get(url2)
            browser_get(browser, url2)
            html = browser.page_source
            doc = pq(html)
            yzm = doc('.confirm .confirm-content .yzm').text()
            while yzm:
                # browser.close()
                # login(browser)
                proxy = get_proxy()
                options.add_argument('--proxy-server=http://' + proxy)
                browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe", options=options)
                browser.get(url2)
                html = browser.page_source
                doc = pq(html)
                yzm = doc('.confirm #validCaptchaForm .confirm-content .yzm').text()
                print(yzm)
            row = row + 1
            print('爬取的excel行数row  ===', row)
            row = get_ins(row, doc)
            print('============================================================================================')
        if pages and page_urls:
            for p in pages:
                page.append(p.text())
            for url in page_urls:
                page_url.append(url.find('a').attr('href'))
            dic = dict(zip(page, page_url))
            if '>' in dic:
                wait_yzm()
                # browser.get('http://drugs.medlive.cn/' + dic['>'])
                browser_get(browser, url='http://drugs.medlive.cn/' + dic['>'])
                html = browser.page_source
                doc = pq(html)
                yzm = doc('.confirm .confirm-content .yzm').text()
                while yzm:
                    # browser.close()
                    # login(browser)
                    proxy = get_proxy()
                    options.add_argument('--proxy-server=http://' + proxy)
                    browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe", options=options)
                    browser.get('http://drugs.medlive.cn/' + dic['>'])
                    html = browser.page_source
                    doc = pq(html)
                    yzm = doc('.confirm #validCaptchaForm .confirm-content .yzm').text()
                    print(yzm)
                # while yzm:
                #     wait_yzm()
                #     # browser.get('http://drugs.medlive.cn/' + dic['>'])
                #     browser_get(browser, url='http://drugs.medlive.cn/' + dic['>'])
                #     html = browser.page_source
                #     doc = pq(html)
                #     yzm = doc('.confirm #validCaptchaForm .confirm-content .yzm').text()
                #     print(yzm)
                row = get_instruction(browser, html, row, i)
        return row
    else:
        print(f"搜索出与输入的药品名相关的药物有1种：{medince_name}")
        row = row + 1
        print('row  ===', row)
        row = get_ins(row, doc)
        print('============================================================================================')
        return row


def get_ins(row, doc):
    titles = doc('.info-content .info-left .title').items()
    informations = doc('.info-content .info-left .more-infomation').items()
    tit = []
    inf = []
    for title in titles:
        # print("名称-------", pq(title)('.title').text())
        tit.append(pq(title)('.title').text())
    for information in informations:
        # print("内容-------", pq(information)('.more-infomation').text())
        inf.append(pq(information)('.more-infomation').text())
    # print(dict(zip(tit, inf)))
    for k, v in dict(zip(tit, inf)).items():
        if k[:-1] in titltname and k[:-1] == titltname[titltname.index(k[:-1])]:
            for j in range(1, len(list(zip(tit, inf))) + 1):
                currentSheet.cell(row=row, column=titltname.index(k[:-1]) + 1, value=v)
        if k == "药品名称：":
            print(v)
    return row


def main():
    # chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])  # 以键值对的形式加入参数,隐藏selenium防一定程度的反爬
    # chrome_options.add_argument('headless')
    # browser = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)
    browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe")
    login(browser)

    # 直接爬取所有药品
    browser.get('http://drugs.medlive.cn/drugref/drugCateIndex.do')
    html = browser.page_source
    doc = pq(html)
    # print(html)
    urls = doc('.three-table .table2 table tr td span a').items()

    # # 分配url分爬
    # urls = []
    # f = open(r'西药url1.txt', 'r', encoding='utf-8')
    # while True:
    #     url = f.readline()
    #     if url == '':
    #         break
    #     urls.append(url[:-1])

    row = 1
    i = 0
    for url in urls:
        drug_url = 'http://drugs.medlive.cn/' + url.attr('href')
        # drug_url = url
        print("正在获取网页数据......")
        print(drug_url)
        browser_get(browser, drug_url)
        html = browser.page_source
        doc = pq(html)
        yzm = doc('.confirm .confirm-content .yzm').text()
        while yzm:
            # browser.close()
            # login(browser)
            proxy = get_proxy()
            options.add_argument('--proxy-server=https://' + proxy)
            browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe", options=options)
            browser.get(drug_url)
            html = browser.page_source
            doc = pq(html)
            yzm = doc('.confirm #validCaptchaForm .confirm-content .yzm').text()
            print(yzm)
            sleep(1)
            submit = doc('.left.nav_search .logo-nav .nav.clearfix .drugInfo').text()
            if not submit:
                yzm = True
                browser.close()

        # while yzm:
        #     wait_yzm()
        #     browser_get(browser, drug_url)
        #     html = browser.page_source
        #     doc = pq(html)
        #     yzm = doc('.confirm #validCaptchaForm .confirm-content .yzm').text()
        #     print(yzm)
        row = get_instruction(browser, html, row, i)
        wb.save("药品说明书-西药url1.xlsx")


if __name__ == '__main__':
    main()
