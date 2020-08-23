from openpyxl import load_workbook
import os


f = open(r'.\result_data\爬取的结果为空白的药品.txt', 'w', encoding='utf-8')
src_path = r"药品说明书(全).xlsx"
path = os.path.abspath(src_path)
print("《{src_path}》的路径是：{path}".format(src_path=src_path, path=path))
file = load_workbook(path)
active_sheet = file["Sheet1"]
for i in range(2, active_sheet.max_row):
    drugName = active_sheet.cell(row=i, column=2).value
    if drugName is None and active_sheet.cell(row=i, column=1).value:
        f.write(active_sheet.cell(row=i, column=1).value + '\n')
f.close()
while True:
    file1 = load_workbook(path)
    active_sheet1 = file1["Sheet1"]
    space = []
    for i in range(1, active_sheet1.max_row):
        drugName = active_sheet1.cell(row=i, column=2).value
        if drugName is None:
            space.append(drugName)
            active_sheet1.delete_rows(i)
    file1.save(path)
    if len(space) == 0:
        break
