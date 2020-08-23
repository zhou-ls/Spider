from openpyxl import load_workbook


def find_sentence(en1, en2, drug_interaction):
    drug_interaction = drug_interaction.split('\n')
    all_sentence = []  # 存一个药品说明书的所有句子 先分行后分句 以句号分句
    sentence = []  # 包含所有药物对句子的列表
    for line in drug_interaction:
        line = line.split('。')
        for i in line:
            if i != '':
                i = i.replace(' ', '')  # 去除句子中的空格
                all_sentence.append(i)
    for sen in all_sentence:
        if (en1 in sen or "本品" in sen) and en2 in sen:
            sentence.append(sen.replace('本品', en1))
    return sentence


file = load_workbook(r'.\excel_data\药品说明书(爬取).xlsx')
ac_sheet = file["Sheet1"]
f = open(r'.\origin_data\all_drug_name.txt', 'r', encoding='utf-8')
f1 = open(r'.\origin_data\exchange.txt', 'w', encoding='utf-8')
all_drug = []
while True:
    drug_name = f.readline()
    if drug_name == '':
        break
    # print(type(drug_name[:-1]))
    all_drug.append(drug_name[:-1])

print(len(all_drug))
for i in range(1, ac_sheet.max_row):
    ins = ac_sheet.cell(row=i, column=16).value
    en1 = ac_sheet.cell(row=i, column=1).value.replace(' ', '')
    for j in range(len(all_drug)):
        drug = all_drug[j].replace(' ', '')
        if drug is not None and drug != '无' and ins is not None and en1 != drug and drug in ins:
            sentences = find_sentence(en1, drug, ins)
            for all_sen in sentences:
                # #  句子实体对数据
                # if en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("避免" in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("禁止" in all_sen.replace(' ', '') or ("禁用" in all_sen.replace(' ', '') or "禁忌"in all_sen.replace(' ', '') or "终止"in all_sen.replace(' ', ''))):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "注意" in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("建议" in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "不推荐" in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "谨慎" in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "监测" in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("慎用" in all_sen.replace(' ', '') or "慎重" in all_sen.replace(' ', '')or "宜" in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # # if drug[i][0] in all_sen and drug[i][1] in all_sen and ("减少" in drug[i][2] or "减小"in drug[i][2]):
                # #     f.write('临床建议' + '      ' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and (("增加" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', '')) or ("调整" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', '')) or ("加大" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', ''))):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "二者合用效果更好"in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "其他方式避孕"in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("不得"in all_sen.replace(' ', '') or "不应"in all_sen.replace(' ', '') or "密切观察"in all_sen.replace(' ', '') or "小时"in all_sen.replace(' ', '') or "日"in all_sen.replace(' ', '')) or "天"in all_sen.replace(' ', ''):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("暂停"in all_sen.replace(' ', '') or "更换"in all_sen.replace(' ', '') or "间隔"in all_sen.replace(' ', '') or "改用"in all_sen.replace(' ', '') or "换用"in all_sen.replace(' ', '') or "不要"in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and (("药效"in all_sen.replace(' ', '') or "毒性"in all_sen.replace(' ', '')or "暴露"in all_sen.replace(' ', '')or "效果"in all_sen.replace(' ', '')or "反应"in all_sen.replace(' ', '') )and ("增高"in all_sen.replace(' ', '') or "提高"in all_sen.replace(' ', '')or "升高"in all_sen.replace(' ', '')or "下降"in all_sen.replace(' ', '')or "增大"in all_sen.replace(' ', '')or "减少"in all_sen.replace(' ', '')or "减弱"in all_sen.replace(' ', '')or "减小"in all_sen.replace(' ', '')or "降低"in all_sen.replace(' ', '')or "加重"in all_sen.replace(' ', '')or "减轻"in all_sen.replace(' ', ''))):
                # #     f1.write(en1 + '\t' + drug + '\t' + '相互作用结果' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("发生"in all_sen.replace(' ', '') or"引起"in all_sen.replace(' ', '') or"导致"in all_sen.replace(' ', '') or"药效"in all_sen.replace(' ', '') or "疗效"in all_sen.replace(' ', '') or "毒性"in all_sen.replace(' ', '')or "暴露"in all_sen.replace(' ', '')or "效果"in all_sen.replace(' ', '')or "反应"in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '相互作用结果' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("血药浓度"in all_sen.replace(' ', '') or "时间缩短"in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '相互作用结果' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("增强"in all_sen.replace(' ', '') and "作用"in all_sen.replace(' ', '')):
                #     f1.write(en1 + '\t' + drug + '\t' + '相互作用结果' + '\t' + all_sen.replace(' ', '') + '\n')
                # else:
                #     f1.write(en1 + '\t' + drug + '\t' + '相互作用机制' + '\t' + all_sen.replace(' ', '') + '\n')

                # # bert数据
                # if en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("避免" in all_sen.replace(' ', '')):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("禁止" in all_sen.replace(' ', '') or ("禁用" in all_sen.replace(' ', '') or "禁忌"in all_sen.replace(' ', '') or "终止"in all_sen.replace(' ', ''))):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "注意" in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("建议" in all_sen.replace(' ', '')):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "不推荐" in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "谨慎" in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "监测" in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("慎用" in all_sen.replace(' ', '') or "慎重" in all_sen.replace(' ', '')or "宜" in all_sen.replace(' ', '')):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # # if drug[i][0] in all_sen and drug[i][1] in all_sen and ("减少" in drug[i][2] or "减小"in drug[i][2]):
                # #     f.write('临床建议' + '      ' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and (("增加" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', '')) or ("调整" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', '')) or ("加大" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', ''))):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "二者合用效果更好"in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "其他方式避孕"in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("不得"in all_sen.replace(' ', '') or "不应"in all_sen.replace(' ', '') or "密切观察"in all_sen.replace(' ', '') or "小时"in all_sen.replace(' ', '') or "日"in all_sen.replace(' ', '')) or "天"in all_sen.replace(' ', ''):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("暂停"in all_sen.replace(' ', '') or "更换"in all_sen.replace(' ', '') or "间隔"in all_sen.replace(' ', '') or "改用"in all_sen.replace(' ', '') or "换用"in all_sen.replace(' ', '') or "不要"in all_sen.replace(' ', '')):
                #     f1.write('临床建议' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and (("药效"in all_sen.replace(' ', '') or "毒副作用"in all_sen.replace(' ', '')or "效果"in all_sen.replace(' ', '')or "反应"in all_sen.replace(' ', '') )and ("增高"in all_sen.replace(' ', '') or "提高"in all_sen.replace(' ', '')or "增大"in all_sen.replace(' ', '')or "减少"in all_sen.replace(' ', '')or "减弱"in all_sen.replace(' ', '')or "减小"in all_sen.replace(' ', '')or "降低"in all_sen.replace(' ', '')or "加重"in all_sen.replace(' ', '')or "减轻"in all_sen.replace(' ', ''))):
                #     f1.write('相互作用结果' + '\t' + all_sen.replace(' ', '') + '\n')
                # else:
                #     f1.write('相互作用机制' + '\t' + all_sen.replace(' ', '') + '\n')
                #
                # 核对的句子实体对
                if en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("避免" in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("禁止" in all_sen.replace(' ', '') or ("禁用" in all_sen.replace(' ', '') or "禁忌"in all_sen.replace(' ', '') or "终止"in all_sen.replace(' ', ''))):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "注意" in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("建议" in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "不推荐" in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "谨慎" in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "监测" in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("慎用" in all_sen.replace(' ', '') or "慎重" in all_sen.replace(' ', '')or "宜" in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                # if drug[i][0] in all_sen and drug[i][1] in all_sen and ("减少" in drug[i][2] or "减小"in drug[i][2]):
                #     f.write('临床建议' + '      ' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and (("增加" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', '')) or ("调整" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', '')) or ("加大" in all_sen.replace(' ', '') and "剂量"in all_sen.replace(' ', ''))):
                    f1.write(en1 + '\t' + drug + '\t'+ '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "二者合用效果更好"in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and "其他方式避孕"in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("不得"in all_sen.replace(' ', '') or "不应"in all_sen.replace(' ', '') or "密切观察"in all_sen.replace(' ', '') or "小时"in all_sen.replace(' ', '') or "日"in all_sen.replace(' ', '')) or "天"in all_sen.replace(' ', ''):
                    f1.write(en1 + '\t' + drug + '\t'+ '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("暂停"in all_sen.replace(' ', '') or "更换"in all_sen.replace(' ', '') or "间隔"in all_sen.replace(' ', '') or "改用"in all_sen.replace(' ', '') or "换用"in all_sen.replace(' ', '') or "不要"in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                # elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and (("药效"in all_sen.replace(' ', '') or "毒副作用"in all_sen.replace(' ', '')or "效果"in all_sen.replace(' ', '')or "反应"in all_sen.replace(' ', '') )and ("增高"in all_sen.replace(' ', '') or "提高"in all_sen.replace(' ', '')or "增大"in all_sen.replace(' ', '')or "减少"in all_sen.replace(' ', '')or "减弱"in all_sen.replace(' ', '')or "减小"in all_sen.replace(' ', '')or "降低"in all_sen.replace(' ', '')or "加重"in all_sen.replace(' ', '')or "减轻"in all_sen.replace(' ', ''))):
                #     f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("发生"in all_sen.replace(' ', '') or"引起"in all_sen.replace(' ', '') or"导致"in all_sen.replace(' ', '') or"药效"in all_sen.replace(' ', '') or "疗效"in all_sen.replace(' ', '') or "毒性"in all_sen.replace(' ', '')or "暴露"in all_sen.replace(' ', '')or "效果"in all_sen.replace(' ', '')or "反应"in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("血药浓度"in all_sen.replace(' ', '') or "时间缩短"in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                elif en1 in all_sen.replace(' ', '') and drug in all_sen.replace(' ', '') and ("增强"in all_sen.replace(' ', '') and "作用"in all_sen.replace(' ', '')):
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
                else:
                    f1.write(en1 + '\t' + drug + '\t' + '\t' + all_sen.replace(' ', '') + '\n')
f1.close()
f2 = open(r'.\origin_data\exchange.txt', 'r', encoding='utf-8')
f3 = open(r'.\result_data\实体对+句子+三分类.txt', 'w', encoding='utf-8')
lines = []
while True:
    line = f2.readline()
    if line == '':
        break
    # print(type(drug_name[:-1]))
    lines.append(line[:-1])
# print(lines)
lines = list(set(lines))
print(len(lines))
for n in lines:
    f3.write(n + '\n')
f3.close()
