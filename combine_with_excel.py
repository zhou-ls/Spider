#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020/3/23 12:50
# @Author : zls
# @Site : 
# @File : combine_with_excel.py
# @Software: PyCharm
from openpyxl import Workbook, load_workbook


wb = Workbook()
wb.create_sheet(index=0, title="Sheet" + str(1))
file1 = load_workbook(r'.\excel_data\药品说明书(爬取).xlsx')
file3 = load_workbook(r'.\excel_data\药品说明书.xlsx')
file2 = load_workbook(r'.\excel_data\药品说明书-除空.xlsx')
ac_sheet1 = file1["Sheet1"]
ac_sheet2 = file2["Sheet1"]
ac_sheet3 = file3["Sheet1"]
currentSheet = wb.active
# print(ac_sheet1.cell(row=13873, column=1).value)
i = 1
a = 0
b = 1
c = 1
while True:
    a = a + 1
    print(i)
    if ac_sheet1.cell(row=a, column=1).value is not None:
        for j in range(1, 33):
            currentSheet.cell(row=i, column=j, value=ac_sheet1.cell(row=a, column=j).value)
        i = i + 1
        if i % 1000 == 0:
            wb.save("药品说明书（爬取结果）.xlsx")
    else:
        break
print("读取《药品说明书(爬取).xlsx》完毕")
while True:
    b = b + 1
    print(i)
    if ac_sheet2.cell(row=b, column=1).value is not None:
        for j in range(1, 33):
            currentSheet.cell(row=i, column=j, value=ac_sheet2.cell(row=b, column=j).value)
        # wb.save("药品说明书（爬取结果）.xlsx")
        i = i + 1
        if i % 1000 == 0:
            wb.save("药品说明书（爬取结果）.xlsx")
    else:
        break
print("读取《药品说明书-除空.xlsx》完毕")
while True:
    c = c + 1
    print(i)
    if ac_sheet3.cell(row=c, column=1).value is not None:
        for j in range(1, 33):
            currentSheet.cell(row=i, column=j, value=ac_sheet3.cell(row=c, column=j).value)
        # wb.save("药品说明书（爬取结果）.xlsx")
        i = i + 1
        if i % 50 == 0:
            wb.save("药品说明书（爬取结果）.xlsx")
    else:
        break
print("读取《药品说明书.xlsx》完毕")
