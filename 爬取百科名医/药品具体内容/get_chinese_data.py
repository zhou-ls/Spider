# -*- coding: utf-8 -*-
# @Time : 2020/8/11 20:28
# @Author : zls
# @File : get_chinese_data.py
# @Software: PyCharm
import re
from time import sleep
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook
from random import randint

# chrome_options = webdriver.ChromeOptions()
# chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])  # 以键值对的形式加入参数,隐藏selenium防一定程度的反爬
# chrome_options.add_argument('headless')
# browser = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)
browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe")
wait = WebDriverWait(browser, 10)

title = ['通用名称', '名称', '本药是什么？', '功效与作用', '用药前需要注意哪些事项？', '如何使用？', '服药期间我应该避免什么？', '药物毒性与不良反应', '如何选购？', '含有本药的中成药有哪些？',
         '与其他药功效一样吗？', '用药误区'
                       '其他注意事项']

# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active
for i in range(1, len(title) + 1):
    currentSheet.cell(row=1, column=i, value=title[i - 1])


# # 打开excel文档追加爬取
# wb = load_workbook("名医百科药品.xlsx")
# currentSheet = wb["Sheet1"]


# 如果网络超时，则重新爬取数据
def browser_get(url):
    try:
        browser.get(url)
    except TimeoutException:
        print('网络超时，正在尝试重新爬取数据......')
        browser_get(url)


def search(row):
    browser_get('https://www.baikemy.com/medicine/medicineListIndex?pageType=1')
    submit = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, '#content > div.first-tab > div.content-tab > div:nth-child(3)')))
    submit.click()
    sleep(5)
    html = browser.page_source
    # print(html)
    doc = pq(html)
    drug_urls = doc('#safe-area #medicine-content  #content .first-tab div .info-box.chinese .info-left .content-list a').items()
    # print(drug_urls)
    for url in drug_urls:
        row += 1
        print(url.attr('href'))
        browser_get('https://www.baikemy.com' + url.attr('href'))
        html = browser.page_source
        doc = pq(html)
        # 当前爬取的药品名称
        detail_name = doc('.main_wrap .content_wrap .content_left .detail_name_wrap .detail_name').text()
        currentSheet.cell(row=row, column=1, value=detail_name)
        # 当前爬取的药品通用名
        name_info = doc('.main_wrap .content_wrap .content_left .name_info').text()
        print(name_info)
        currentSheet.cell(row=row, column=2, value=name_info)
        # 具体内容
        content = doc('.main_wrap .content_wrap .content_left .content .directory_flag').items()
        num = 3
        for con in content:
            # p = con.find('p').items()
            title = con.text()
            # print(title)
            # print('--------------')
            title_content = con.siblings().text()
            title_content = title + '\n' + title_content
            currentSheet.cell(row=row, column=num, value=title_content)
            num += 1
            # print(title_content)
            # print('------------------------------------------------------------------------')


if __name__ == '__main__':
    row = 1
    search(row)
    wb.save("名医百科中药.xlsx")
