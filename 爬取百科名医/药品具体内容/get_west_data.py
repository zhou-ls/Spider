# -*- coding: utf-8 -*-
# @Time : 2020/8/11 17:26
# @Author : zls
# @File : get_west_data.py
# @Software: PyCharm
import re
from time import sleep
from pyquery import PyQuery as pq
import requests
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# chrome_options = webdriver.ChromeOptions()
# chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])  # 以键值对的形式加入参数,隐藏selenium防一定程度的反爬
# chrome_options.add_argument('headless')
# browser = webdriver.Chrome("D:\chrom\Chrome-bin\chromedriver.exe", chrome_options=chrome_options)
browser = webdriver.Chrome(executable_path=r"D:\chrom\Chrome-bin\chromedriver.exe")
wait = WebDriverWait(browser, 10)

title = ['通用名称', '名称', '是什么药？', '有哪些用途？', '有哪些制剂和规格？', '用药前须知', '我该如何用药？', '用药期间注意事项', '可能有哪些不良反应？', '相互作用', '妊娠期和哺乳期能否使用该药？',
         '如何合理使用该药？']

# 创建excel文档
wb = Workbook()
wb.create_sheet(index=0, title="Sheet1")
currentSheet = wb.active
for i in range(1, len(title) + 1):
    currentSheet.cell(row=1, column=i, value=title[i - 1])

# # 打开excel文档追加爬取
# wb = load_workbook("名医百科药品.xlsx")
# currentSheet = wb["Sheet1"]

# 维持会话
session = requests.Session()
session.keep_alive = False


def get_html(url):
    try:
        response = session.get(url)
        # 解决中文乱码
        response.encoding = response.apparent_encoding
        html = response.text
        return html
    except ConnectionError:
        print('网络超时，正在尝试重新爬取数据......')
        return get_html(url)


# 如果网络超时，则重新爬取数据
def browser_get(url):
    try:
        browser.get(url)
    except TimeoutException:
        print('网络超时，正在尝试重新爬取数据......')
        browser_get(url)


# 如果要读取多个药品，则反复调用search()
def search(row):
    browser_get('https://www.baikemy.com/medicine/medicineListIndex?pageType=1')
    submit = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, '#content > div.first-tab > div.content-tab > div:nth-child(2)')))
    submit.click()
    sleep(5)
    html = browser.page_source
    # print(html)
    doc = pq(html)
    drug_urls = doc('.info-box.western .info-left .content-list a').items()
    for url in drug_urls:
        row += 1
        print(url.attr('href'))
        html = get_html('https://www.baikemy.com' + url.attr('href'))
        doc = pq(html)
        # 当前爬取的药品名称
        detail_name = doc('.main_wrap .content_wrap .content_left .detail_name_wrap .detail_name').text()
        currentSheet.cell(row=row, column=1, value=detail_name)
        # 当前爬取的药品通用名
        name_info = doc('.main_wrap .content_wrap .content_left .name_info').text()
        print(name_info)
        currentSheet.cell(row=row, column=2, value=name_info)
        # 具体内容
        content = doc('.main_wrap .content_wrap .content_left .content .directory_flag').items()
        num = 3
        for con in content:
            # p = con.find('p').items()
            # title = con.text()
            # print(title)
            # print('--------------')
            title_content = con.siblings().text()
            currentSheet.cell(row=row, column=num, value=title_content)
            num += 1
            # print(title_content)
            # print('------------------------------------------------------------------------')


if __name__ == '__main__':
    row = 1
    search(row)
    wb.save("名医百科西药.xlsx")
