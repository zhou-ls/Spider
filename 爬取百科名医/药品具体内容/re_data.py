# -*- coding: utf-8 -*-
# @Time : 2020/8/12 8:21
# @Author : zls
# @File : re_data.py
# @Software: PyCharm
from openpyxl import load_workbook, Workbook
import re


wb1 = Workbook()
wb1.create_sheet(index=0, title="Sheet1")
currentSheet1 = wb1.active

wb2 = Workbook()
wb2.create_sheet(index=0, title="Sheet1")
currentSheet2 = wb2.active
